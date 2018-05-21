import openpyxl.cell

wb = openpyxl.load_workbook('example.xlsx')

a = wb.sheetnames
b = wb[a[0]]

print(a, b, b.title)

for i in range(1, 8, 2):
    print(b.cell(row=i, column=2).value)
    print(b.cell.column_index_from_string())

print()